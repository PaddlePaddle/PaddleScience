# coding=utf-8

import pandas as pd
from openpyxl import load_workbook

# 读取 Excel 文件
file_path = '工作簿4.xlsx'
wb = load_workbook(file_path, data_only=True)
sheet = wb.active  # 或者使用 wb['SheetName'] 来指定特定的 sheet

# 创建一个空的 DataFrame 来存储数据
df = pd.DataFrame()
key_list = []
value_list = []
# 遍历所有行和列
i = 0
for row in sheet.iter_rows(min_row=2, values_only=False):  # min_row=2 从第二行开始
    print("row: ", i)
    i = i+1
    row_data = []
    for cell in row:
        if cell.hyperlink:  # 检查单元格是否有超链接
            row_data.append(cell.hyperlink)  # 添加超链接的 URL
            cell.value = cell.value.replace(" - 飞桨AI Studio星河社区", '')
            cell.value = cell.value.replace("﻿", '')
            print(cell.value) #打印cell name
            print(cell.hyperlink.display)
            key_list.append(cell.value)
            value_list.append(cell.hyperlink.display)
        else:
            row_data.append(cell.value)  # 添加单元格的值

n = len(value_list)

with open('notebooks.md', 'w', encoding='utf-8') as file:
    # 遍历字典的 key 和 value
    for i in range(n):
        # 按照 [key](value) 的格式写入文件
        file.write(f'[{key_list[i]}]({value_list[i]})\n\n')

# 打印 DataFrame
print(df)